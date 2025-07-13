from openpyxl import Workbook, load_workbook
import os

def salvar_dados_em_excel(data, temperatura, umidade):
    arquivo = "dados.xlsx"

    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "Temperatura", "Umidade"])

    ws.append([data, temperatura, umidade])
    wb.save(arquivo)
